# -*coding: UTF-8 -*-
"""
----------------------------------
@Author: 般若
@time：2019/8/22 18:11
@File：ReadExcel
-----------------------------------
"""
from openpyxl import load_workbook
from config.constants import cases_file_path,config_file_path
from common.HandleConfig import HandleConfig

hc = HandleConfig(config_file_path)
prefix = hc.get_value("url", "prefix")
result_col = hc.get_int("case_file", "result_col")
actual_col = hc.get_int("case_file", "actual_col")

class ReadExcel:
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname


    def getExcelCases(self):
        wb = load_workbook(self.filename)
        if self.sheetname == None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        head_data = tuple(ws.iter_rows(values_only=True))[0]
        cell_data = tuple(ws.iter_rows(min_row=2, values_only=True))

        list = []
        for data in cell_data:
            list.append(dict(zip(head_data, data)))


        return list

    def write_result(self, row, actual, result):
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        ws.cell(row=row, column=actual_col, value=actual)
        ws.cell(row=row, column=result_col, value=result)
        wb.save(self.filename)


if __name__ == '__main__':
    re = ReadExcel(cases_file_path, "music")
    casess = re.getExcelCases()
    print(casess)
    tt = casess[0]["datas"]
    print(tt,type(tt))